#!/usr/bin/env python3
"""Regenerate the JSON fixtures used by verify.js from the two workbooks.
Run: python3 make_fixtures.py   (requires openpyxl)
Produces _sample.json (BCA_Trial_data.xlsx) and _test2.json (BCA_test_data_2.xlsx)."""
import openpyxl, json, sys
def dump(xlsx, out):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    json.dump({"sheet": ws.title, "rows": rows}, open(out, "w"), default=str)
    print(f"{xlsx} -> {out} ({len(rows)} rows, sheet '{ws.title}')")
dump("BCA_Trial_data.xlsx", "_sample.json")
dump("BCA_test_data_2.xlsx", "_test2.json")
